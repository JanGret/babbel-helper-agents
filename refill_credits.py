"""
Babbel: Intensive Credits verwalten (hinzufuegen oder entfernen).

Loggt sich im Babbel-Portal ein, navigiert zur Intensive Credits Seite,
sucht einen Nutzer per E-Mail und fuegt Credits hinzu oder entfernt sie.

Nutzung:
    py refill_credits.py --email user@example.com             # 10 Credits hinzufuegen (Standard)
    py refill_credits.py --email user@example.com --credits 20  # 20 Credits hinzufuegen
    py refill_credits.py --email user@example.com --remove      # 10 Credits entfernen
    py refill_credits.py --email user@example.com --remove --credits 5  # 5 Credits entfernen

Voraussetzungen:
    - .env Datei mit BABBEL_EMAIL und BABBEL_PASSWORD
    - py -m pip install playwright python-dotenv
    - py -m playwright install chromium
"""

import argparse
import io
import os
import sys
import time
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

from babbel_utils import read_pool_info, print_pool_info

load_dotenv()
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

BABBEL_BASE_URL = "https://my.babbel.com"
INTENSIVE_URL = f"{BABBEL_BASE_URL}/en/organizations/jobcloud/intensivecredits"
SESSION_DIR = Path(__file__).parent / ".session"


# ---------------------------------------------------------------------------
# Browser Automation
# ---------------------------------------------------------------------------

def login(page):
    """Login to Babbel. Skips if session is still valid."""
    email = os.getenv("BABBEL_EMAIL")
    password = os.getenv("BABBEL_PASSWORD")
    if not email or not password:
        print("FEHLER: BABBEL_EMAIL und BABBEL_PASSWORD muessen in .env gesetzt sein.")
        sys.exit(1)

    print(f"Login als {email}...")
    login_url = (
        f"{BABBEL_BASE_URL}/de/authentication/login/email"
        "?return_to=https%3A%2F%2Fmy.babbel.com%2Fde%2Forganizations%2Fjobcloud%2Fusers"
    )
    page.goto(login_url, wait_until="domcontentloaded")

    # Bereits eingeloggt?
    try:
        page.wait_for_url("**/organizations/**", timeout=15000)
        print("  Bereits eingeloggt (gespeicherte Session).")
        return
    except PlaywrightTimeout:
        pass

    # E-Mail eingeben
    email_input = page.locator('input[type="email"], input[name="email"], input[id*="email"]')
    email_input.first.wait_for(state="visible", timeout=15000)
    email_input.first.fill(email)
    page.locator('button:has-text("Weiter"), button:has-text("Continue"), [type="submit"]').first.click()
    print("  E-Mail eingegeben...")

    # Passwort eingeben
    pw_input = page.locator('input[type="password"]')
    pw_input.first.wait_for(state="visible", timeout=15000)
    pw_input.first.fill(password)
    page.locator('button:has-text("Einloggen"), button:has-text("Log in"), button:has-text("Weiter"), [type="submit"]').first.click()
    print("  Passwort eingegeben...")

    # Captcha-Check
    try:
        page.wait_for_url("**/organizations/**", timeout=10000)
    except PlaywrightTimeout:
        print("\n  *** CAPTCHA — bitte im Browser loesen (max 5 Min) ***")
        try:
            page.wait_for_url("**/organizations/**", timeout=300000)
        except PlaywrightTimeout:
            if "authentication" in page.url:
                print("FEHLER: Login fehlgeschlagen.")
                sys.exit(1)

    print("  Login erfolgreich.")


def _read_remaining_from_row(page, row_element):
    """Liest den 'Remaining' Wert aus einer Tabellenzeile.

    Ermittelt den Spaltenindex von 'Remaining' im Header und liest
    den entsprechenden Wert aus der Datenzeile.

    Args:
        page: Playwright page object
        row_element: Locator fuer die Tabellenzeile des Nutzers

    Returns:
        int oder None falls nicht auslesbar
    """
    import re

    try:
        # Ansatz 1: Spaltenindex ueber Header ermitteln
        headers = page.locator('thead th, thead td, [role="columnheader"]')
        header_count = headers.count()

        remaining_col_idx = None
        for i in range(header_count):
            header_text = (headers.nth(i).text_content() or "").strip().lower()
            if "remaining" in header_text:
                remaining_col_idx = i
                break

        if remaining_col_idx is not None:
            # Lese die entsprechende Zelle in der Datenzeile
            cells = row_element.locator('td')
            if cells.count() > remaining_col_idx:
                cell_text = (cells.nth(remaining_col_idx).text_content() or "").strip()
                match = re.search(r'(\d+)', cell_text)
                if match:
                    return int(match.group(1))

        # Ansatz 2: Fallback - alle Zellen durchsuchen, die nach "remaining" aussehen
        # (falls die Tabelle keine thead hat oder der Index nicht stimmt)
        row_text = row_element.text_content() or ""
        # Suche nach Mustern wie "Remaining: 5" oder einfach die letzte Zahl
        cells = row_element.locator('td')
        cell_count = cells.count()
        for i in range(cell_count):
            cell_text = (cells.nth(i).text_content() or "").strip()
            # Pruefe ob die Spalte ueberschrift "Remaining" hat
            if header_count > i:
                h_text = (headers.nth(i).text_content() or "").strip().lower()
                if "remaining" in h_text:
                    match = re.search(r'(\d+)', cell_text)
                    if match:
                        return int(match.group(1))

        return None

    except Exception:
        return None


def update_excel(user_email, credits_amount, remove=False):
    """Aktualisiert die SharePoint-Excel nach Credit-Aenderung.

    - Spalte F (Credits Intensive 2026): credits_amount (positiv bei Add, negativ bei Remove)
    - Spalte G (Credit Restock 2026): Heutiges Datum im Format MM/DD/YYYY

    Args:
        user_email: E-Mail-Adresse des Nutzers
        credits_amount: Anzahl Credits (immer positiv)
        remove: True wenn Credits entfernt wurden

    Returns:
        True bei Erfolg, False bei Fehler
    """
    excel_path = os.getenv("SHAREPOINT_EXCEL_PATH")
    if not excel_path:
        print("  WARNUNG: SHAREPOINT_EXCEL_PATH nicht in .env gesetzt.")
        return False

    excel_path = Path(excel_path)
    if not excel_path.exists():
        print(f"  WARNUNG: Excel-Datei nicht gefunden: {excel_path}")
        return False

    try:
        wb = load_workbook(excel_path)
        ws = wb.active

        # Nutzer per E-Mail in Spalte B suchen
        for row_idx in range(2, ws.max_row + 1):
            email_cell = ws.cell(row=row_idx, column=2).value
            if email_cell and email_cell.strip().lower() == user_email.lower():
                # Spalte F (6): Credits (positiv bei Add, negativ bei Remove)
                value = -credits_amount if remove else credits_amount
                ws.cell(row=row_idx, column=6).value = value

                # Spalte G (7): Restock-Datum
                ws.cell(row=row_idx, column=7).value = datetime.now()
                ws.cell(row=row_idx, column=7).number_format = "MM/DD/YYYY"

                wb.save(excel_path)
                wb.close()
                return True

        wb.close()
        print(f"  WARNUNG: Nutzer '{user_email}' nicht in Excel gefunden.")
        return False

    except Exception as e:
        print(f"  WARNUNG: Excel-Update fehlgeschlagen: {e}")
        return False


def manage_credits(page, user_email, credits_amount, remove=False, force=False):
    """Navigate to intensive credits page, find user and add/remove credits."""
    action = "entfernen" if remove else "hinzufuegen"
    print(f"\nNavigiere zu Intensive Credits Seite...")
    page.goto(INTENSIVE_URL)
    try:
        page.wait_for_load_state("networkidle", timeout=30000)
    except PlaywrightTimeout:
        pass
    time.sleep(3)

    # Suchfeld finden und E-Mail eingeben
    print(f"  Suche Nutzer: {user_email}...")
    search_input = page.locator(
        'input[type="search"], '
        'input[placeholder*="Search"], '
        'input[placeholder*="search"], '
        'input[placeholder*="Suche"], '
        'input[placeholder*="mail"], '
        'input[data-class*="search"], '
        '[role="searchbox"]'
    )
    search_input.first.wait_for(state="visible", timeout=15000)
    search_input.first.fill(user_email)
    time.sleep(2)

    # Warte bis Suchergebnisse geladen sind (Tabelle aktualisiert sich)
    print("  Warte auf Suchergebnisse...")
    user_row = page.locator(f'tr:has-text("{user_email}"), [data-class*="row"]:has-text("{user_email}")')
    try:
        user_row.first.wait_for(state="visible", timeout=15000)
    except PlaywrightTimeout:
        print(f"  FEHLER: Nutzer '{user_email}' nicht in der Tabelle gefunden.")
        sys.exit(1)

    # Remaining Credits aus der Tabellenzeile auslesen
    remaining = _read_remaining_from_row(page, user_row.first)
    if remaining is not None:
        print(f"  Aktuelle Credits remaining: {remaining}")
    else:
        print(f"  WARNUNG: Konnte 'Remaining' nicht aus der Tabelle auslesen.")

    # Pruefung: Nur beim Hinzufuegen und wenn nicht --force
    if not remove and not force and remaining is not None:
        new_total = remaining + credits_amount

        # Hard-Cap: Wuerde Maximum 20 ueberschreiten
        if new_total > 20:
            max_possible = 20 - remaining
            print(f"\n  ABBRUCH: Nutzer hat bereits {remaining} Credits remaining.")
            print(f"  Mit +{credits_amount} waeren es {new_total} — Maximum ist 20.")
            if max_possible > 0:
                print(f"  Maximal noch {max_possible} Credits moeglich.")
                print(f"  → Erneut ausfuehren mit: --credits {max_possible}")
            else:
                print(f"  Nutzer hat bereits das Maximum erreicht.")
            print(f"  → Um die Pruefung zu ueberspringen: --force")
            sys.exit(3)

        # Soft-Warnung: User hat bereits >= 5 remaining
        if remaining >= 5:
            diff_to_10 = 10 - remaining
            print(f"\n  HINWEIS: Nutzer hat bereits {remaining} Credits remaining.")
            if diff_to_10 > 0:
                print(f"  Empfehlung: {diff_to_10} Credits auffuellen (auf 10 total).")
                print(f"  Alternativ: {credits_amount} Credits hinzufuegen (auf {new_total} total).")
            else:
                print(f"  Nutzer hat bereits >= 10 Credits. Kein Auffuellen noetig.")
                print(f"  Falls gewuenscht: {credits_amount} Credits hinzufuegen (auf {new_total} total).")
            print(f"  → Erneut ausfuehren mit angepasstem --credits oder --force")
            sys.exit(2)

    # Nutzer per Checkbox/Klick auswaehlen (aktiviert die Buttons)
    print(f"  Waehle Nutzer aus...")
    checkbox = user_row.first.locator('input[type="checkbox"], [role="checkbox"]')
    if checkbox.count() > 0:
        checkbox.first.click(timeout=5000)
    else:
        # Fallback: Klicke auf die Zeile selbst
        user_row.first.click(timeout=5000)
    time.sleep(2)

    # Button klicken: "Add Credits" oder "Remove Credits"
    if remove:
        print(f"  Oeffne 'Remove Credits' Popup...")
        btn = page.locator('[data-class="remove-intensivecredits-cta"]')
        if btn.count() == 0:
            btn = page.locator(
                'button:has-text("Remove Credits"), '
                'a:has-text("Remove Credits"), '
                '[aria-label*="Remove Credits"]'
            )
    else:
        print(f"  Oeffne 'Add Credits' Popup...")
        btn = page.locator('[data-class="add-intensivecredits-cta"]')
        if btn.count() == 0:
            btn = page.locator(
                'button:has-text("Add Credits"), '
                'a:has-text("Add Credits"), '
                '[aria-label*="Add Credits"]'
            )
    btn.first.click(timeout=10000)
    time.sleep(2)

    # Popup: Credits-Anzahl eingeben
    print(f"  Setze Credits auf {credits_amount}...")

    # Finde das Nummern-Eingabefeld im Popup/Modal
    credits_input = page.locator(
        '[role="dialog"] input[type="number"], '
        '[role="dialog"] input[type="text"], '
        '[class*="modal"] input[type="number"], '
        '[class*="modal"] input[type="text"], '
        '[class*="popup"] input[type="number"], '
        '[class*="popup"] input[type="text"], '
        '[data-class*="credits"] input'
    )
    try:
        credits_input.first.wait_for(state="visible", timeout=10000)
    except PlaywrightTimeout:
        # Fallback: Jedes sichtbare Nummern-Eingabefeld
        credits_input = page.locator('input[type="number"]:visible')
        credits_input.first.wait_for(state="visible", timeout=5000)

    # Feld leeren und Wert eingeben
    credits_input.first.clear()
    credits_input.first.fill(str(credits_amount))
    time.sleep(1)

    # "Confirm" Button klicken
    print("  Bestaetigung klicken...")
    confirm_btn = page.locator(
        '[role="dialog"] button:has-text("Confirm"), '
        '[class*="modal"] button:has-text("Confirm"), '
        'button:has-text("Confirm"), '
        'button:has-text("Bestaetigen"), '
        'button:has-text("Bestätigen")'
    )
    confirm_btn.first.click(timeout=10000)
    time.sleep(3)

    # Pruefe auf Meldung (Erfolg oder Fehler)
    action_past = "entfernt" if remove else "hinzugefuegt"
    alert = page.locator('[role="alert"]:visible')
    if alert.count() > 0:
        alert_text = alert.first.text_content() or ""
        if "success" in alert_text.lower() or "changed" in alert_text.lower():
            # Erfolgsmeldung vom Portal - Pool-Info auslesen
            time.sleep(2)  # Warten bis Seite aktualisiert
            pool_info = read_pool_info(page)
            print(f"\n{'=' * 50}")
            print(f"FERTIG: {credits_amount} Credits fuer {user_email} {action_past}.")
            print(f"  Portal-Meldung: {alert_text.strip()}")
            print_pool_info(pool_info)
            # Excel aktualisieren
            if update_excel(user_email, credits_amount, remove):
                print(f"  Excel: Credits und Restock-Datum eingetragen.")
            print(f"{'=' * 50}")
            return
        else:
            print(f"  FEHLER vom Portal: {alert_text}")
            sys.exit(1)

    # Pruefe auf explizite Fehlermeldungen
    error = page.locator('[class*="error"]:visible')
    if error.count() > 0:
        error_text = error.first.text_content()
        print(f"  FEHLER vom Portal: {error_text}")
        sys.exit(1)

    # Kein Alert gefunden, aber auch kein Fehler - Pool-Info trotzdem auslesen
    time.sleep(2)
    pool_info = read_pool_info(page)
    print(f"\n{'=' * 50}")
    print(f"FERTIG: {credits_amount} Credits fuer {user_email} {action_past}.")
    print_pool_info(pool_info)
    # Excel aktualisieren
    if update_excel(user_email, credits_amount, remove):
        print(f"  Excel: Credits und Restock-Datum eingetragen.")
    print(f"{'=' * 50}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Babbel: Intensive Credits verwalten")
    parser.add_argument("--email", required=True,
                        help="E-Mail-Adresse des Nutzers")
    parser.add_argument("--credits", type=int, default=10,
                        help="Anzahl Credits (Standard: 10)")
    parser.add_argument("--remove", action="store_true",
                        help="Credits entfernen statt hinzufuegen")
    parser.add_argument("--force", action="store_true",
                        help="Credits-Pruefung ueberspringen (Soft-Warnung und Hard-Cap ignorieren)")
    parser.add_argument("--headless", action="store_true",
                        help="Browser ohne Fenster ausfuehren")
    args = parser.parse_args()

    if args.credits <= 0:
        print("FEHLER: Credits muessen > 0 sein.")
        sys.exit(1)

    action = "entfernen" if args.remove else "hinzufuegen"
    print(f"Modus: {args.credits} Credits {action} fuer {args.email}")
    if args.force:
        print("  (--force: Credits-Pruefung wird uebersprungen)")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=args.headless)
        SESSION_DIR.mkdir(exist_ok=True)
        context = browser.new_context(
            storage_state=str(SESSION_DIR / "state.json") if (SESSION_DIR / "state.json").exists() else None,
            viewport={"width": 1440, "height": 900},
        )
        page = context.new_page()

        try:
            login(page)
            context.storage_state(path=str(SESSION_DIR / "state.json"))
            manage_credits(page, args.email, args.credits, remove=args.remove, force=args.force)

        except Exception as e:
            # Screenshot bei Fehler
            data_dir = Path(__file__).parent / "daten"
            data_dir.mkdir(exist_ok=True)
            page.screenshot(path=str(data_dir / "refill_error_screenshot.png"))
            print(f"\nFEHLER: {e}")
            print(f"Screenshot: {data_dir / 'refill_error_screenshot.png'}")
            sys.exit(1)
        finally:
            browser.close()


if __name__ == "__main__":
    main()
