"""
Babbel: Neuen Nutzer einladen.

Loggt sich im Babbel-Portal ein, navigiert zur Users-Seite,
klickt "Invite New Member" und fuellt das Einladungs-Popup aus.

Nutzung:
    py invite_user.py --email new@example.com --name "Max Mustermann" --lang en
    py invite_user.py --email new@example.com --name "Marie Dupont" --plan intensive --lang fr
    py invite_user.py --email new@example.com --name "Hans Mueller" --plan professional --lang de

Voraussetzungen:
    - .env Datei mit BABBEL_EMAIL und BABBEL_PASSWORD
    - py -m pip install playwright python-dotenv openpyxl
    - py -m playwright install chromium
"""

import argparse
import csv
import io
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

load_dotenv()
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

BABBEL_BASE_URL = "https://my.babbel.com"
USERS_URL = f"{BABBEL_BASE_URL}/en/organizations/jobcloud/users"
SESSION_DIR = Path(__file__).parent / ".session"


# ---------------------------------------------------------------------------
# Browser Automation
# ---------------------------------------------------------------------------

def login(page):
    """Login to Babbel. Skips if session is still valid."""
    email = os.getenv("BABBEL_EMAIL")
    password = os.getenv("BABBEL_PASSWORD")
    if not email or not password:
        print("FEHLER: BABBEL_EMAIL und BABBEL_PASSWORD muessen in .env gesetzt sein.")
        sys.exit(1)

    print(f"Login als {email}...")
    login_url = (
        f"{BABBEL_BASE_URL}/de/authentication/login/email"
        "?return_to=https%3A%2F%2Fmy.babbel.com%2Fde%2Forganizations%2Fjobcloud%2Fusers"
    )
    page.goto(login_url, wait_until="domcontentloaded")

    # Bereits eingeloggt?
    try:
        page.wait_for_url("**/organizations/**", timeout=15000)
        print("  Bereits eingeloggt (gespeicherte Session).")
        return
    except PlaywrightTimeout:
        pass

    # E-Mail eingeben
    email_input = page.locator('input[type="email"], input[name="email"], input[id*="email"]')
    email_input.first.wait_for(state="visible", timeout=15000)
    email_input.first.fill(email)
    page.locator('button:has-text("Weiter"), button:has-text("Continue"), [type="submit"]').first.click()
    print("  E-Mail eingegeben...")

    # Passwort eingeben
    pw_input = page.locator('input[type="password"]')
    pw_input.first.wait_for(state="visible", timeout=15000)
    pw_input.first.fill(password)
    page.locator('button:has-text("Einloggen"), button:has-text("Log in"), button:has-text("Weiter"), [type="submit"]').first.click()
    print("  Passwort eingegeben...")

    # Captcha-Check
    try:
        page.wait_for_url("**/organizations/**", timeout=10000)
    except PlaywrightTimeout:
        print("\n  *** CAPTCHA — bitte im Browser loesen (max 5 Min) ***")
        try:
            page.wait_for_url("**/organizations/**", timeout=300000)
        except PlaywrightTimeout:
            if "authentication" in page.url:
                print("FEHLER: Login fehlgeschlagen.")
                sys.exit(1)

    print("  Login erfolgreich.")


def _select_dropdown_option(page, dropdown_locator, option_text):
    """Oeffnet ein Custom-Dropdown und waehlt eine Option per Text.

    Funktioniert mit gaengigen Custom-Select-Implementierungen
    (MUI, styled-components, etc.).

    Args:
        page: Playwright page
        dropdown_locator: Locator fuer das Dropdown-Element
        option_text: Text der zu waehlenden Option

    Returns:
        True bei Erfolg, False bei Fehler
    """
    try:
        # Klicke auf das Dropdown um es zu oeffnen
        dropdown_locator.click(timeout=5000)
        time.sleep(1)

        # Suche nach der Option in verschiedenen moeglichen Containern
        option = page.locator(
            f'li:has-text("{option_text}"), '
            f'[role="option"]:has-text("{option_text}"), '
            f'[role="listbox"] >> text="{option_text}", '
            f'[class*="option"]:has-text("{option_text}"), '
            f'[class*="menu-item"]:has-text("{option_text}")'
        )
        option.first.click(timeout=5000)
        time.sleep(1)
        return True
    except (PlaywrightTimeout, Exception) as e:
        print(f"    WARNUNG: Dropdown-Option '{option_text}' nicht gefunden: {e}")
        return False


def _extract_name_from_email(email):
    """Extrahiert Vor- und Nachname aus vorname.nachname@... Format.

    Args:
        email: E-Mail-Adresse (z.B. florian.vallet@jobcloud.ch)

    Returns:
        str (z.B. "Florian Vallet") oder None falls nicht extrahierbar
    """
    try:
        local_part = email.split("@")[0]  # z.B. "florian.vallet"
        parts = local_part.split(".")     # z.B. ["florian", "vallet"]
        if len(parts) >= 2:
            return " ".join(p.capitalize() for p in parts)
        return None
    except Exception:
        return None


def _check_duplicate_email(user_email):
    """Prueft ob die E-Mail bereits in memberships.csv existiert.

    Args:
        user_email: E-Mail-Adresse die geprueft werden soll

    Returns:
        None falls nicht gefunden, oder dict mit Nutzer-Info falls gefunden
    """
    csv_path = Path(__file__).parent / "daten" / "memberships.csv"
    if not csv_path.exists():
        return None  # CSV nicht vorhanden, Check ueberspringen

    try:
        with open(csv_path, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                # Versuche verschiedene Spaltennamen fuer E-Mail
                email = (row.get("Email") or row.get("email") or row.get("E-Mail") or row.get("E-mail") or "").strip().lower()
                if email == user_email.strip().lower():
                    return row
    except Exception:
        return None

    return None


def _read_member_count(page):
    """Liest 'X members total' von der Users-Seite.

    Returns:
        int (Anzahl Members) oder None falls nicht auslesbar
    """
    try:
        page_text = page.text_content("body") or ""
        match = re.search(r"(\d+)\s*members?\s*total", page_text, re.IGNORECASE)
        if match:
            return int(match.group(1))
        return None
    except Exception:
        return None


def _count_intensive_in_excel():
    """Zaehlt Intensive-Nutzer in der SharePoint-Excel (Spalte E = 'x').

    Returns:
        int (Anzahl Intensive-Nutzer) oder None falls nicht auslesbar
    """
    excel_path = os.getenv("SHAREPOINT_EXCEL_PATH")
    if not excel_path:
        return None

    excel_path = Path(excel_path)
    if not excel_path.exists():
        return None

    try:
        wb = load_workbook(excel_path, read_only=True)
        ws = wb.active
        count = 0
        for row_idx in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=5).value
            if cell_value and str(cell_value).strip().lower() in ("x", "true", "1"):
                count += 1
        wb.close()
        return count
    except Exception:
        return None


def invite_user(page, user_email, plan, lang, credits_amount):
    """Navigate to users page and invite a new member."""
    print(f"\nNavigiere zu Users Seite...")
    page.goto(USERS_URL)
    try:
        page.wait_for_load_state("networkidle", timeout=30000)
    except PlaywrightTimeout:
        pass
    time.sleep(3)

    # Slot-Check: Gesamt-Slots pruefen
    # "X members total" beinhaltet 3 Admin-Accounts die keine Nutzer-Slots belegen
    ADMIN_COUNT = 3
    MAX_SLOTS = 100
    slot_warning = False
    member_count = _read_member_count(page)
    if member_count is not None:
        user_slots_used = member_count - ADMIN_COUNT
        slots_remaining = MAX_SLOTS - user_slots_used
        print(f"  Aktuelle Belegung: {user_slots_used}/{MAX_SLOTS} Nutzer-Slots ({slots_remaining} frei)")
        if user_slots_used >= MAX_SLOTS:
            slot_warning = True
            print(f"\n  WARNUNG: Account hat {user_slots_used}/{MAX_SLOTS} Nutzer-Slots belegt!")
            print(f"  ({member_count} members total inkl. {ADMIN_COUNT} Admins)")
            print(f"  Einladung wird trotzdem durchgefuehrt (kurzfristige Ueberschreitung moeglich).")
            print(f"  EMPFEHLUNG: Inaktive Nutzer pruefen mit: py consolidate.py\n")
    else:
        print(f"  Slot-Belegung konnte nicht ausgelesen werden.")

    # Slot-Check: Intensive-Limit pruefen (nur bei --plan intensive)
    intensive_warning = False
    if plan == "intensive":
        intensive_count = _count_intensive_in_excel()
        if intensive_count is not None:
            print(f"  Intensive-Nutzer aktiv: {intensive_count}/30 (Budget-Kalkulation)")
            if intensive_count >= 30:
                intensive_warning = True
                print(f"\n  WARNUNG: Bereits {intensive_count} Intensive-Nutzer aktiv (Budget kalkuliert fuer max. 30).")
                print(f"  Einladung wird trotzdem durchgefuehrt.\n")

    # "Invite New Member" Button klicken
    print("  Klicke 'Invite New Member'...")
    invite_btn = page.locator(
        'button:has-text("Invite New Member"), '
        'button:has-text("Invite new member"), '
        '[data-class*="invite"], '
        'a:has-text("Invite New Member")'
    )
    invite_btn.first.click(timeout=10000)
    time.sleep(2)

    # Popup: E-Mail-Adresse eingeben
    print(f"  E-Mail eingeben: {user_email}...")
    email_input = page.locator(
        '[role="dialog"] input[type="email"], '
        '[role="dialog"] input[type="text"], '
        '[class*="modal"] input[type="email"], '
        '[class*="modal"] input[type="text"], '
        '[class*="invite"] input[type="email"], '
        '[class*="invite"] input[type="text"]'
    )
    # Das erste Input-Feld im Dialog ist typischerweise das E-Mail-Feld
    email_input.first.wait_for(state="visible", timeout=10000)
    email_input.first.fill(user_email)
    time.sleep(1)

    # Einladungssprache auswaehlen (Radix UI Dropdown Menu)
    print(f"  Einladungssprache: {lang}...")
    lang_labels = {
        "en": "English",
        "de": "Deutsch",
        "fr": "Français",
        "es": "Español",
        "it": "Italiano",
        "pt": "Português",
        "pl": "Polski",
        "sv": "Svenska",
        "uk": "Українська",
    }
    lang_label = lang_labels.get(lang, lang)

    # Klicke den Language-Trigger-Button (hat "Email invitation language" als Text)
    lang_trigger = page.locator(
        '[role="dialog"] button:has-text("Email invitation language"), '
        '[role="dialog"] button:has-text("invitation language"), '
        'button:has-text("Email invitation language")'
    )
    if lang_trigger.count() > 0:
        lang_trigger.first.click(timeout=5000)
        time.sleep(1)
        # Waehle die Sprache per data-class Attribut
        lang_option = page.locator(f'[data-class="selectlanguage-invitation-{lang}"]')
        if lang_option.count() > 0:
            lang_option.first.click(timeout=5000)
        else:
            # Fallback: per role="menuitem" und Text
            lang_option = page.locator(f'[role="menuitem"]:has-text("{lang_label}")')
            lang_option.first.click(timeout=5000)
    else:
        print(f"    WARNUNG: Language-Trigger nicht gefunden. Standardsprache wird verwendet.")
    time.sleep(1)

    # Plan auswaehlen (natives <select> mit data-class="selectplan-invitation-standard")
    print(f"  Plan: {plan}...")
    plan_label = "Professional" if plan == "professional" else "Intensive"

    plan_dropdown = page.locator(
        '[data-class="selectplan-invitation-standard"], '
        'select[aria-label="Plan"], '
        '[role="dialog"] select[data-class*="selectplan"]'
    )

    if plan_dropdown.count() > 0:
        plan_dropdown.first.select_option(label=plan_label)
    else:
        print(f"    WARNUNG: Plan-Dropdown nicht gefunden.")
    time.sleep(2)

    # Falls Intensive: Start-Credits eingeben
    if plan == "intensive":
        print(f"  Start-Credits: {credits_amount}...")
        credits_input = page.locator(
            '[role="dialog"] input[type="number"], '
            '[role="dialog"] input[placeholder*="credit" i], '
            '[role="dialog"] input[placeholder*="Credit" i], '
            '[role="dialog"] [data-class*="credit"] input'
        )
        try:
            credits_input.first.wait_for(state="visible", timeout=5000)
            credits_input.first.clear()
            credits_input.first.fill(str(credits_amount))
            time.sleep(1)
        except PlaywrightTimeout:
            print(f"    WARNUNG: Credits-Feld nicht gefunden. Wird moeglicherweise nicht angezeigt.")

    # "Invite New Members" Button klicken
    print("  Einladung absenden...")
    confirm_btn = page.locator(
        '[data-class="add-invitations-modal"], '
        '[role="dialog"] button:has-text("Invite New Members"), '
        '[role="dialog"] button:has-text("Invite new members"), '
        'button:has-text("Invite New Members")'
    )
    confirm_btn.first.click(timeout=10000)
    time.sleep(3)

    # Pruefe auf Meldung (Erfolg oder Fehler)
    alert = page.locator('[role="alert"]:visible')
    if alert.count() > 0:
        alert_text = alert.first.text_content() or ""
        if "success" in alert_text.lower() or "invited" in alert_text.lower() or "sent" in alert_text.lower():
            print(f"\n{'=' * 50}")
            print(f"FERTIG: Einladung an {user_email} gesendet.")
            print(f"  Plan: {plan_label}")
            if plan == "intensive":
                print(f"  Start-Credits: {credits_amount}")
            print(f"  Sprache: {lang_label}")
            print(f"  Portal-Meldung: {alert_text.strip()}")
            print(f"{'=' * 50}")
            return True
        else:
            print(f"  FEHLER vom Portal: {alert_text}")
            return False

    # Pruefe auf explizite Fehlermeldungen
    error = page.locator('[class*="error"]:visible, [class*="Error"]:visible')
    if error.count() > 0:
        error_text = error.first.text_content()
        print(f"  FEHLER vom Portal: {error_text}")
        return False

    # Kein Alert, kein Fehler - vermutlich erfolgreich
    print(f"\n{'=' * 50}")
    print(f"FERTIG: Einladung an {user_email} gesendet.")
    print(f"  Plan: {plan_label}")
    if plan == "intensive":
        print(f"  Start-Credits: {credits_amount}")
    print(f"  Sprache: {lang_label}")
    print(f"{'=' * 50}")
    return True


# ---------------------------------------------------------------------------
# Excel Update
# ---------------------------------------------------------------------------

def add_user_to_excel(user_name, user_email, plan, credits_amount):
    """Traegt einen neuen Nutzer in die SharePoint-Excel ein.

    Neue Zeile direkt nach dem letzten aktiven Eintrag:
    - Spalte A: Name
    - Spalte B: E-Mail
    - Spalte C: Entry (heutiges Datum)
    - Spalte D: Professional ("x" wenn professional)
    - Spalte E: Intensive 2026 ("x" wenn intensive)
    - Spalte F: Credits Intensive 2026 (10 wenn intensive)
    - Spalte G: Credit Restock 2026 (Datum wenn intensive)
    - Spalte L: Status ("Invited")

    Returns:
        True bei Erfolg, False bei Fehler
    """
    excel_path = os.getenv("SHAREPOINT_EXCEL_PATH")
    if not excel_path:
        print("  WARNUNG: SHAREPOINT_EXCEL_PATH nicht in .env gesetzt.")
        return False

    excel_path = Path(excel_path)
    if not excel_path.exists():
        print(f"  WARNUNG: Excel-Datei nicht gefunden: {excel_path}")
        return False

    try:
        wb = load_workbook(excel_path)
        ws = wb.active

        # Erste freie Zeile finden: Letzte Zeile mit E-Mail in Spalte B + 1
        next_row = 2  # Mindestens Zeile 2 (nach Header)
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=2).value:
                next_row = row_idx + 1

        # Spalte A: Name
        ws.cell(row=next_row, column=1).value = user_name

        # Spalte B: E-Mail
        ws.cell(row=next_row, column=2).value = user_email

        # Spalte C: Entry (heutiges Datum)
        ws.cell(row=next_row, column=3).value = datetime.now()
        ws.cell(row=next_row, column=3).number_format = "MM/DD/YYYY"

        # Spalte D: Professional
        if plan == "professional":
            ws.cell(row=next_row, column=4).value = "x"

        # Spalte E: Intensive 2026
        if plan == "intensive":
            ws.cell(row=next_row, column=5).value = "x"

            # Spalte F: Credits Intensive (2026)
            ws.cell(row=next_row, column=6).value = credits_amount

            # Spalte G: Credit Restock (2026)
            ws.cell(row=next_row, column=7).value = datetime.now()
            ws.cell(row=next_row, column=7).number_format = "MM/DD/YYYY"

        # Spalte L (12): Status
        ws.cell(row=next_row, column=12).value = "Invited"

        wb.save(excel_path)
        wb.close()
        return True

    except Exception as e:
        print(f"  WARNUNG: Excel-Update fehlgeschlagen: {e}")
        return False


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Babbel: Neuen Nutzer einladen")
    parser.add_argument("--email", required=True,
                        help="E-Mail-Adresse des neuen Nutzers")
    parser.add_argument("--name", default=None,
                        help="Vor- und Nachname (optional: wird aus vorname.nachname@... extrahiert)")
    parser.add_argument("--plan", choices=["professional", "intensive"], default="professional",
                        help="Plan-Typ (Standard: professional)")
    parser.add_argument("--lang", required=True,
                        help="Einladungssprache (en, de, fr, es, it, pt)")
    parser.add_argument("--credits", type=int, default=10,
                        help="Start-Credits bei Intensive (Standard: 10)")
    parser.add_argument("--headless", action="store_true",
                        help="Browser ohne Fenster ausfuehren")
    args = parser.parse_args()

    # Name: Explizit angegeben oder aus E-Mail extrahieren
    if args.name:
        user_name = args.name
    else:
        user_name = _extract_name_from_email(args.email)
        if not user_name:
            print(f"FEHLER: Name konnte nicht aus E-Mail '{args.email}' extrahiert werden.")
            print(f"  Bitte --name angeben (z.B. --name \"Max Mustermann\").")
            print(f"  Automatische Extraktion funktioniert nur bei vorname.nachname@... Format.")
            sys.exit(1)
        print(f"  Name aus E-Mail extrahiert: {user_name}")

    plan_label = "Professional" if args.plan == "professional" else "Intensive"
    print(f"Einladung: {user_name} ({args.email})")
    print(f"  Plan: {plan_label}, Sprache: {args.lang}")
    if args.plan == "intensive":
        print(f"  Start-Credits: {args.credits}")

    # Duplikat-Pruefung: Existiert die E-Mail bereits?
    existing = _check_duplicate_email(args.email)
    if existing:
        name = existing.get("Name") or existing.get("name") or existing.get("Full Name") or "Unbekannt"
        print(f"\nABBRUCH: Nutzer mit E-Mail {args.email} existiert bereits im Babbel-Account.")
        print(f"  Name: {name}")
        print(f"  Kein Invite durchgefuehrt.")
        print(f"  Falls die CSV veraltet ist: Vorher `py export_babbel.py --tab memberships` ausfuehren.")
        sys.exit(1)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=args.headless)
        SESSION_DIR.mkdir(exist_ok=True)
        context = browser.new_context(
            storage_state=str(SESSION_DIR / "state.json") if (SESSION_DIR / "state.json").exists() else None,
            viewport={"width": 1440, "height": 900},
        )
        page = context.new_page()

        try:
            login(page)
            context.storage_state(path=str(SESSION_DIR / "state.json"))

            success = invite_user(page, args.email, args.plan, args.lang, args.credits)

            if success:
                # Excel aktualisieren
                print("\n  Excel aktualisieren...")
                if add_user_to_excel(user_name, args.email, args.plan, args.credits):
                    print(f"  Excel: Neuer Nutzer eingetragen.")
                else:
                    print(f"  WARNUNG: Excel-Eintrag fehlgeschlagen.")
            else:
                print("\n  Einladung fehlgeschlagen. Excel wird NICHT aktualisiert.")
                sys.exit(1)

        except Exception as e:
            # Screenshot bei Fehler
            data_dir = Path(__file__).parent / "daten"
            data_dir.mkdir(exist_ok=True)
            page.screenshot(path=str(data_dir / "invite_error_screenshot.png"))
            print(f"\nFEHLER: {e}")
            print(f"Screenshot: {data_dir / 'invite_error_screenshot.png'}")
            sys.exit(1)
        finally:
            browser.close()


if __name__ == "__main__":
    main()
