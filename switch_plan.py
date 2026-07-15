"""
Babbel: Plan wechseln (Professional → Intensive).

Stellt einen bestehenden Professional-Nutzer auf den Intensive-Plan um
und fuellt anschliessend 10 Start-Credits nach.

Nutzung:
    py switch_plan.py --email user@example.com             # Switch + 10 Credits
    py switch_plan.py --email user@example.com --credits 20  # Switch + 20 Credits

Voraussetzungen:
    - .env Datei mit BABBEL_EMAIL und BABBEL_PASSWORD
    - py -m pip install playwright python-dotenv openpyxl
    - py -m playwright install chromium
"""

import argparse
import csv
import io
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

from babbel_utils import read_pool_info, print_pool_info

load_dotenv()
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

BABBEL_BASE_URL = "https://my.babbel.com"
USERS_URL = f"{BABBEL_BASE_URL}/en/organizations/jobcloud/users"
INTENSIVE_URL = f"{BABBEL_BASE_URL}/en/organizations/jobcloud/intensivecredits"
SESSION_DIR = Path(__file__).parent / ".session"
DATA_DIR = Path(__file__).parent / "daten"


# ---------------------------------------------------------------------------
# Browser Automation
# ---------------------------------------------------------------------------

def login(page):
    """Login to Babbel. Skips if session is still valid."""
    email = os.getenv("BABBEL_EMAIL")
    password = os.getenv("BABBEL_PASSWORD")
    if not email or not password:
        print("FEHLER: BABBEL_EMAIL und BABBEL_PASSWORD muessen in .env gesetzt sein.")
        sys.exit(1)

    print(f"Login als {email}...")
    login_url = (
        f"{BABBEL_BASE_URL}/de/authentication/login/email"
        "?return_to=https%3A%2F%2Fmy.babbel.com%2Fde%2Forganizations%2Fjobcloud%2Fusers"
    )
    page.goto(login_url, wait_until="domcontentloaded")

    # Bereits eingeloggt?
    try:
        page.wait_for_url("**/organizations/**", timeout=15000)
        print("  Bereits eingeloggt (gespeicherte Session).")
        return
    except PlaywrightTimeout:
        pass

    # E-Mail eingeben
    email_input = page.locator('input[type="email"], input[name="email"], input[id*="email"]')
    email_input.first.wait_for(state="visible", timeout=15000)
    email_input.first.fill(email)
    page.locator('button:has-text("Weiter"), button:has-text("Continue"), [type="submit"]').first.click()
    print("  E-Mail eingegeben...")

    # Passwort eingeben
    pw_input = page.locator('input[type="password"]')
    pw_input.first.wait_for(state="visible", timeout=15000)
    pw_input.first.fill(password)
    page.locator('button:has-text("Einloggen"), button:has-text("Log in"), button:has-text("Weiter"), [type="submit"]').first.click()
    print("  Passwort eingegeben...")

    # Captcha-Check
    try:
        page.wait_for_url("**/organizations/**", timeout=10000)
    except PlaywrightTimeout:
        print("\n  *** CAPTCHA — bitte im Browser loesen (max 5 Min) ***")
        try:
            page.wait_for_url("**/organizations/**", timeout=300000)
        except PlaywrightTimeout:
            if "authentication" in page.url:
                print("FEHLER: Login fehlgeschlagen.")
                sys.exit(1)

    print("  Login erfolgreich.")


# ---------------------------------------------------------------------------
# Pre-Checks
# ---------------------------------------------------------------------------

def _check_user_exists(email):
    """Prueft ob der User in memberships.csv existiert.

    Returns:
        True wenn gefunden, False wenn nicht
    """
    csv_path = DATA_DIR / "memberships.csv"
    if not csv_path.exists():
        return None  # Kann nicht pruefen

    try:
        with open(csv_path, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                row_email = (row.get("Email") or row.get("email") or "").strip().lower()
                if row_email == email.strip().lower():
                    return True
    except Exception:
        return None
    return False


def _check_already_intensive(email):
    """Prueft ob der User bereits in intensive_credits.csv auftaucht.

    Returns:
        True wenn bereits Intensive, False wenn nicht, None wenn nicht pruefbar
    """
    csv_path = DATA_DIR / "intensive_credits.csv"
    if not csv_path.exists():
        return None

    try:
        with open(csv_path, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                row_email = (row.get("Email") or row.get("email") or "").strip().lower()
                if row_email == email.strip().lower():
                    return True
    except Exception:
        return None
    return False


# ---------------------------------------------------------------------------
# Plan Switch
# ---------------------------------------------------------------------------

def switch_plan(page, user_email):
    """Wechselt den Plan eines Users von Professional auf Intensive (Private Classes).

    Args:
        page: Playwright page object
        user_email: E-Mail-Adresse des Nutzers

    Returns:
        True bei Erfolg, False bei Fehler
    """
    print(f"\nNavigiere zu Users Seite...")
    page.goto(USERS_URL)
    try:
        page.wait_for_load_state("networkidle", timeout=30000)
    except PlaywrightTimeout:
        pass
    time.sleep(3)

    # User suchen
    print(f"  Suche Nutzer: {user_email}...")
    search_input = page.locator(
        'input[type="search"], '
        'input[placeholder*="Search"], '
        'input[placeholder*="search"], '
        '[role="searchbox"]'
    )
    search_input.first.wait_for(state="visible", timeout=15000)
    search_input.first.fill(user_email)
    time.sleep(2)

    # User in Tabelle finden
    print("  Warte auf Suchergebnisse...")
    user_row = page.locator(f'tr:has-text("{user_email}"), [data-class*="row"]:has-text("{user_email}")')
    try:
        user_row.first.wait_for(state="visible", timeout=15000)
    except PlaywrightTimeout:
        print(f"  FEHLER: Nutzer '{user_email}' nicht in der Tabelle gefunden.")
        return False

    # User per Checkbox auswaehlen
    print(f"  Nutzer gefunden. Waehle Nutzer aus...")
    checkbox = user_row.first.locator('input[type="checkbox"], [role="checkbox"]')
    if checkbox.count() > 0:
        checkbox.first.click(timeout=5000)
    else:
        user_row.first.click(timeout=5000)
    time.sleep(2)

    # "Switch Plan" Button klicken
    print(f"  Klicke 'Switch Plan'...")
    switch_btn = page.locator(
        'button:has-text("Switch Plan"), '
        'button:has-text("Switch plan"), '
        '[data-class*="switch-plan"], '
        '[data-class*="switch_plan"]'
    )
    switch_btn.first.click(timeout=10000)
    time.sleep(2)

    # "Private Classes" auswaehlen (Radio/Karte)
    print(f"  Waehle 'Private Classes'...")
    private_classes = page.locator(
        '[role="radio"]:has-text("Private Classes"), '
        '[role="radio"]:has-text("Private classes"), '
        'label:has-text("Private Classes"), '
        '[data-class*="private"], '
        '[class*="card"]:has-text("Private Classes"), '
        'button:has-text("Private Classes"), '
        'div:has-text("Private Classes") >> input[type="radio"]'
    )
    if private_classes.count() > 0:
        private_classes.first.click(timeout=5000)
    else:
        # Fallback: Suche nach Text "Private" in einem klickbaren Element
        fallback = page.locator('[role="dialog"] >> text=Private Classes')
        fallback.first.click(timeout=5000)
    time.sleep(2)

    # "Continue" klicken
    print(f"  Klicke 'Continue'...")
    continue_btn = page.locator(
        'button:has-text("Continue"), '
        'button:has-text("Weiter")'
    )
    continue_btn.first.click(timeout=10000)
    time.sleep(3)

    # "Confirm new Plan" klicken
    print(f"  Klicke 'Confirm new Plan'...")
    confirm_btn = page.locator(
        'button:has-text("Confirm new Plan"), '
        'button:has-text("Confirm new plan"), '
        'button:has-text("Confirm"), '
        '[data-class*="confirm"]'
    )
    confirm_btn.first.click(timeout=10000)
    time.sleep(3)

    # Erfolg pruefen
    alert = page.locator('[role="alert"]:visible')
    if alert.count() > 0:
        alert_text = alert.first.text_content() or ""
        if "success" in alert_text.lower() or "switch" in alert_text.lower() or "changed" in alert_text.lower():
            print(f"  Plan-Switch erfolgreich: {alert_text.strip()}")
            return True
        else:
            print(f"  FEHLER vom Portal: {alert_text}")
            return False

    # Kein Alert, aber auch kein Fehler — vermutlich erfolgreich
    print(f"  Plan-Switch vermutlich erfolgreich (keine Fehlermeldung).")
    return True


# ---------------------------------------------------------------------------
# Credits auffuellen (im selben Browser)
# ---------------------------------------------------------------------------

def refill_credits_after_switch(page, user_email, credits_amount):
    """Fuellt Credits nach dem Plan-Switch auf.

    Navigiert zur Intensive Credits Seite und fuegt Credits hinzu.
    Verwendet den selben Browser/Page der bereits eingeloggt ist.

    Args:
        page: Playwright page object (bereits eingeloggt)
        user_email: E-Mail-Adresse des Nutzers
        credits_amount: Anzahl Credits

    Returns:
        dict mit pool_info oder None bei Fehler
    """
    print(f"\n  Navigiere zu Intensive Credits Seite fuer Credit-Auffuellung...")
    page.goto(INTENSIVE_URL)
    try:
        page.wait_for_load_state("networkidle", timeout=30000)
    except PlaywrightTimeout:
        pass
    time.sleep(3)

    # User suchen
    print(f"  Suche Nutzer: {user_email}...")
    search_input = page.locator(
        'input[type="search"], '
        'input[placeholder*="Search"], '
        'input[placeholder*="search"], '
        '[role="searchbox"]'
    )
    search_input.first.wait_for(state="visible", timeout=15000)
    search_input.first.fill(user_email)
    time.sleep(2)

    # User in Tabelle finden
    user_row = page.locator(f'tr:has-text("{user_email}"), [data-class*="row"]:has-text("{user_email}")')
    try:
        user_row.first.wait_for(state="visible", timeout=15000)
    except PlaywrightTimeout:
        print(f"  WARNUNG: Nutzer nach Plan-Switch nicht in Intensive Credits Tabelle gefunden.")
        print(f"  Moeglicherweise dauert es etwas bis der Switch durchpropagiert.")
        return None

    # User auswaehlen
    checkbox = user_row.first.locator('input[type="checkbox"], [role="checkbox"]')
    if checkbox.count() > 0:
        checkbox.first.click(timeout=5000)
    else:
        user_row.first.click(timeout=5000)
    time.sleep(2)

    # "Add Credits" klicken
    print(f"  Oeffne 'Add Credits' Popup...")
    add_btn = page.locator('[data-class="add-intensivecredits-cta"]')
    if add_btn.count() == 0:
        add_btn = page.locator('button:has-text("Add Credits")')
    add_btn.first.click(timeout=10000)
    time.sleep(2)

    # Credits eingeben
    print(f"  Setze Credits auf {credits_amount}...")
    credits_input = page.locator(
        '[role="dialog"] input[type="number"], '
        '[role="dialog"] input[type="text"], '
        '[class*="modal"] input[type="number"]'
    )
    try:
        credits_input.first.wait_for(state="visible", timeout=10000)
    except PlaywrightTimeout:
        credits_input = page.locator('input[type="number"]:visible')
        credits_input.first.wait_for(state="visible", timeout=5000)

    credits_input.first.clear()
    credits_input.first.fill(str(credits_amount))
    time.sleep(1)

    # Confirm
    print("  Bestaetigung klicken...")
    confirm_btn = page.locator(
        '[role="dialog"] button:has-text("Confirm"), '
        'button:has-text("Confirm")'
    )
    confirm_btn.first.click(timeout=10000)
    time.sleep(3)

    # Erfolg pruefen + Pool-Info
    alert = page.locator('[role="alert"]:visible')
    if alert.count() > 0:
        alert_text = alert.first.text_content() or ""
        if "success" in alert_text.lower() or "changed" in alert_text.lower():
            time.sleep(2)
            pool_info = read_pool_info(page)
            print(f"  Credits: {credits_amount} hinzugefuegt.")
            return pool_info

    # Kein Alert, trotzdem Pool-Info lesen
    time.sleep(2)
    pool_info = read_pool_info(page)
    return pool_info


# ---------------------------------------------------------------------------
# Excel Update
# ---------------------------------------------------------------------------

def update_excel_for_switch(user_email, credits_amount):
    """Aktualisiert die SharePoint-Excel nach Plan-Switch.

    - Spalte E (Intensive 2026): "x"
    - Spalte F (Credits Intensive 2026): credits_amount
    - Spalte G (Credit Restock 2026): Heutiges Datum

    Returns:
        True bei Erfolg, False bei Fehler
    """
    excel_path = os.getenv("SHAREPOINT_EXCEL_PATH")
    if not excel_path:
        print("  WARNUNG: SHAREPOINT_EXCEL_PATH nicht in .env gesetzt.")
        return False

    excel_path = Path(excel_path)
    if not excel_path.exists():
        print(f"  WARNUNG: Excel-Datei nicht gefunden: {excel_path}")
        return False

    try:
        wb = load_workbook(excel_path)
        ws = wb.active

        for row_idx in range(2, ws.max_row + 1):
            email_cell = ws.cell(row=row_idx, column=2).value
            if email_cell and email_cell.strip().lower() == user_email.lower():
                # Spalte E (5): Intensive 2026
                ws.cell(row=row_idx, column=5).value = "x"

                # Spalte F (6): Credits Intensive (2026)
                ws.cell(row=row_idx, column=6).value = credits_amount

                # Spalte G (7): Credit Restock (2026)
                ws.cell(row=row_idx, column=7).value = datetime.now()
                ws.cell(row=row_idx, column=7).number_format = "MM/DD/YYYY"

                wb.save(excel_path)
                wb.close()
                return True

        wb.close()
        print(f"  WARNUNG: Nutzer '{user_email}' nicht in Excel gefunden.")
        return False

    except Exception as e:
        print(f"  WARNUNG: Excel-Update fehlgeschlagen: {e}")
        return False


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Babbel: Plan wechseln (Professional → Intensive)")
    parser.add_argument("--email", required=True,
                        help="E-Mail-Adresse des Nutzers")
    parser.add_argument("--credits", type=int, default=10,
                        help="Start-Credits nach dem Switch (Standard: 10)")
    parser.add_argument("--headless", action="store_true",
                        help="Browser ohne Fenster ausfuehren")
    args = parser.parse_args()

    print(f"Plan-Switch: {args.email}")
    print(f"  Professional → Intensive (Private Classes)")
    print(f"  Start-Credits: {args.credits}")

    # Vorab-Pruefung: Existiert der User?
    exists = _check_user_exists(args.email)
    if exists is False:
        print(f"\nABBRUCH: Nutzer '{args.email}' existiert nicht im Babbel-Account.")
        print(f"  Falls die CSV veraltet ist: Vorher `py export_babbel.py --tab memberships` ausfuehren.")
        sys.exit(1)

    # Vorab-Pruefung: Ist der User bereits Intensive?
    already_intensive = _check_already_intensive(args.email)
    if already_intensive is True:
        print(f"\nABBRUCH: Nutzer '{args.email}' hat bereits den Intensive-Plan.")
        print(f"  Falls Credits nachgefuellt werden sollen: `py refill_credits.py --email {args.email}`")
        sys.exit(2)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=args.headless)
        SESSION_DIR.mkdir(exist_ok=True)
        context = browser.new_context(
            storage_state=str(SESSION_DIR / "state.json") if (SESSION_DIR / "state.json").exists() else None,
            viewport={"width": 1440, "height": 900},
        )
        page = context.new_page()

        try:
            login(page)
            context.storage_state(path=str(SESSION_DIR / "state.json"))

            # Schritt 1: Plan wechseln
            success = switch_plan(page, args.email)

            if not success:
                print("\n  Plan-Switch fehlgeschlagen.")
                sys.exit(1)

            # Schritt 2: Credits auffuellen
            pool_info = refill_credits_after_switch(page, args.email, args.credits)

            # Ergebnis
            print(f"\n{'=' * 50}")
            print(f"FERTIG: Plan-Switch fuer {args.email} abgeschlossen.")
            print(f"  Neuer Plan: Intensive (Private Classes)")
            print(f"  Credits: {args.credits} hinzugefuegt")
            if pool_info:
                print_pool_info(pool_info)

            # Excel aktualisieren
            if update_excel_for_switch(args.email, args.credits):
                print(f"  Excel: Intensive-Flag und Credits eingetragen.")

            print(f"{'=' * 50}")

        except Exception as e:
            DATA_DIR.mkdir(exist_ok=True)
            page.screenshot(path=str(DATA_DIR / "switch_error_screenshot.png"))
            print(f"\nFEHLER: {e}")
            print(f"Screenshot: {DATA_DIR / 'switch_error_screenshot.png'}")
            sys.exit(1)
        finally:
            browser.close()


if __name__ == "__main__":
    main()
