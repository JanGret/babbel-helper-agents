"""
Babbel Sync: SharePoint-Excel mit Daten aus Babbel ergaenzen.

Liest das offizielle Join Date aus memberships.csv und traegt es in die
SharePoint-Excel ein, dort wo das 'Entry'-Feld leer ist.

Nutzung:
    py sync_sharepoint.py

Voraussetzungen:
    - daten/memberships.csv (aktuell, ggf. vorher py export_babbel.py --tab memberships)
    - SharePoint-Excel ueber OneDrive gesynced
    - py -m pip install openpyxl python-dotenv
"""

import csv
import glob
import io
import os
import sys
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

DATA_DIR = Path(__file__).parent / "daten"

SHAREPOINT_EXCEL_PATH = os.getenv(
    "SHAREPOINT_EXCEL_PATH",
    r"C:\Users\JanGretschuskin\Jobcloud\HR - Documents\02_Development\Language courses\Babbel\Babbel User List.xlsx"
)


# ---------------------------------------------------------------------------
# Data Loading
# ---------------------------------------------------------------------------

def load_memberships():
    """Load memberships.csv and return email -> join_date mapping."""
    candidates = [
        DATA_DIR / "memberships.csv",
        *glob.glob(str(DATA_DIR / "memberships*")),
    ]
    filepath = next((f for f in candidates if Path(f).exists()), None)
    if not filepath:
        print("FEHLER: memberships.csv nicht gefunden.")
        print("  Fuehre zuerst aus: py export_babbel.py --tab memberships")
        sys.exit(1)

    print(f"  Memberships: {filepath}")
    data = {}
    with open(filepath, "r", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            email = (row.get("Email") or "").strip().lower()
            join_date_str = (row.get("Join date") or "").strip()
            if email and join_date_str:
                try:
                    # Format: 2024-04-16T06:53:40Z
                    dt = datetime.fromisoformat(join_date_str.replace("Z", "+00:00"))
                    data[email] = dt.replace(tzinfo=None)  # Als naive datetime speichern
                except (ValueError, TypeError):
                    pass
    return data


# ---------------------------------------------------------------------------
# Sync Logic
# ---------------------------------------------------------------------------

def sync_entry_dates():
    """Sync Entry dates from memberships.csv into SharePoint Excel."""
    excel_path = Path(SHAREPOINT_EXCEL_PATH)
    if not excel_path.exists():
        print(f"FEHLER: SharePoint-Excel nicht gefunden: {excel_path}")
        print("  Pruefe ob OneDrive synchronisiert ist.")
        sys.exit(1)

    print(f"  SharePoint-Excel: {excel_path}")
    print()

    # Load memberships
    memberships = load_memberships()
    print(f"  {len(memberships)} Nutzer mit Join Date in memberships.csv")
    print()

    # Open Excel (read-write mode)
    wb = load_workbook(excel_path)
    ws = wb.active

    # Spalten: [0]Name [1]E-mail [2]Entry ...
    # Header-Zeile ueberspringen (row 1)
    updated = []
    skipped_has_entry = 0
    skipped_no_match = 0

    for row_idx in range(2, ws.max_row + 1):
        email_cell = ws.cell(row=row_idx, column=2).value
        entry_cell = ws.cell(row=row_idx, column=3).value
        name_cell = ws.cell(row=row_idx, column=1).value

        if not email_cell:
            continue

        email = str(email_cell).strip().lower()

        # Nur ergaenzen wenn Entry leer ist
        if entry_cell:
            skipped_has_entry += 1
            continue

        # Suche in memberships
        if email in memberships:
            join_date = memberships[email]
            ws.cell(row=row_idx, column=3).value = join_date
            ws.cell(row=row_idx, column=3).number_format = "DD.MM.YYYY"
            updated.append((name_cell or email, email, join_date.strftime("%d.%m.%Y")))
        else:
            skipped_no_match += 1

    # Zusammenfassung
    print(f"Ergebnis:")
    print(f"  Bereits vorhanden (nicht geaendert): {skipped_has_entry}")
    print(f"  Kein Match in memberships.csv: {skipped_no_match}")
    print(f"  Ergaenzt: {len(updated)}")

    if updated:
        print(f"\n  Ergaenzte Entry-Daten:")
        for name, email, date_str in updated:
            print(f"    - {name} ({email}) -> {date_str}")

    if updated:
        wb.save(excel_path)
        print(f"\n  Excel gespeichert: {excel_path.name}")
    else:
        print(f"\n  Keine Aenderungen noetig.")

    wb.close()
    return updated


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Sync: SharePoint-Excel mit Babbel Join Dates ergaenzen")
    print("=" * 60)
    sync_entry_dates()


if __name__ == "__main__":
    main()
