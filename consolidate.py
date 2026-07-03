"""
Babbel Consolidate: Alle Datenquellen zusammenfuehren und Empfehlungen generieren.

Liest Learners, App-Lessons, Intensive Credits, Memberships und SharePoint Excel,
berechnet einen Inactivity-Score pro Nutzer und gibt Empfehlungen aus.

Nutzung:
    py consolidate.py [--context professional|intensive] [--threshold N]

Voraussetzungen:
    - Alle Dateien im daten/ Ordner
    - py -m pip install openpyxl
"""

import argparse
import csv
import io
import glob
import json
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

DATA_DIR = Path(__file__).parent / "daten"
EXPORT_DIR = Path(__file__).parent / "exports"

# Admin-Accounts werden vom Scoring ausgeschlossen
ADMIN_EMAILS = {
    "kyra.fabello@jobcloud.ch",
    "alla.menescardi@jobcloud.ch",
    "jan.gretschuskin@jobcloud.ch",
}

# SharePoint-Excel: Primaer ueber OneDrive Sync, Fallback auf lokale Kopie in daten/
SHAREPOINT_EXCEL_PATH = os.getenv(
    "SHAREPOINT_EXCEL_PATH",
    r"C:\Users\JanGretschuskin\Jobcloud\HR - Documents\02_Development\Language courses\Babbel\Babbel User List.xlsx"
)


# ---------------------------------------------------------------------------
# Data Loaders
# ---------------------------------------------------------------------------

def load_learners():
    """Load Learners CSV: Active Days, Activities, Learning Minutes."""
    # Neue Datei (von export_babbel.py) oder alte Glob-Patterns
    candidates = [
        DATA_DIR / "learners.csv",
        *glob.glob(str(DATA_DIR / "Active_Days*")),
    ]
    filepath = next((f for f in candidates if Path(f).exists()), None)
    if not filepath:
        print("WARNUNG: Learners CSV nicht gefunden.")
        return {}
    print(f"    Quelle: {Path(filepath).name}")
    data = {}
    with open(filepath, "r", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            email = (row.get("E-mail") or row.get("E-Mail") or "").strip().lower()
            if not email:
                continue
            try:
                active_days = int(row.get("Active days", 0) or 0)
            except (ValueError, TypeError):
                active_days = 0
            try:
                learning_min = float(row.get("Learning minutes", 0) or 0)
            except (ValueError, TypeError):
                learning_min = 0.0
            try:
                activities = int(row.get("Activities", 0) or 0)
            except (ValueError, TypeError):
                activities = 0
            data[email] = {
                "name": row.get("Name", "").strip(),
                "active_days": active_days,
                "learning_minutes": round(learning_min, 1),
                "activities": activities,
                "join_date": row.get("Join date", "")[:10],
            }
    return data


def load_app_lessons():
    """Load App-Lessons CSV: First & Latest lesson dates."""
    candidates = [
        DATA_DIR / "app_lessons.csv",
        *glob.glob(str(DATA_DIR / "Displayname*")),
    ]
    filepath = next((f for f in candidates if Path(f).exists()), None)
    if not filepath:
        print("WARNUNG: App-Lessons CSV nicht gefunden.")
        return {}
    print(f"    Quelle: {Path(filepath).name}")
    data = {}
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        headers = next(reader)
        # Spalten: Name, E-mail, Language, Date(first), Level(first), Date(latest), Level(latest)
        for row in reader:
            if len(row) < 7:
                continue
            email = row[1].strip().lower()
            if not email:
                continue
            first_lesson_date = row[3][:10] if row[3] else ""
            latest_lesson_date = row[5][:10] if row[5] else ""
            language = row[2].strip()
            data[email] = {
                "language": language,
                "first_lesson_date": first_lesson_date,
                "latest_lesson_date": latest_lesson_date,
                "latest_lesson_level": row[6].strip() if len(row) > 6 else "",
            }
    return data


def load_intensive_credits():
    """Load Intensive Credits CSV: Total, Used, Remaining credits."""
    candidates = [
        DATA_DIR / "intensive_credits.csv",
        *glob.glob(str(DATA_DIR / "intensive_credits*")),
    ]
    filepath = next((f for f in candidates if Path(f).exists()), None)
    if not filepath:
        print("WARNUNG: Intensive Credits CSV nicht gefunden.")
        return {}
    print(f"    Quelle: {Path(filepath).name}")
    data = {}
    with open(filepath, "r", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            email = (row.get("Email") or "").strip().lower()
            if not email:
                continue
            try:
                remaining = int(row.get("Remaining", 0) or 0)
            except (ValueError, TypeError):
                remaining = 0
            try:
                total = int(row.get("Total", 0) or 0)
            except (ValueError, TypeError):
                total = 0
            try:
                used = int(row.get("Used", 0) or 0)
            except (ValueError, TypeError):
                used = 0
            data[email] = {
                "credits_total": total,
                "credits_used": used,
                "credits_remaining": remaining,
            }
    return data


def load_memberships():
    """Load Memberships CSV: Plan type per user."""
    candidates = [
        DATA_DIR / "memberships.csv",
        *glob.glob(str(DATA_DIR / "memberships*")),
    ]
    filepath = next((f for f in candidates if Path(f).exists()), None)
    if not filepath:
        print("WARNUNG: Memberships CSV nicht gefunden.")
        return {}
    print(f"    Quelle: {Path(filepath).name}")
    data = {}
    with open(filepath, "r", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            email = (row.get("Email") or "").strip().lower()
            if not email:
                continue
            plan_raw = (row.get("Plan") or "").strip()
            if "private-classes" in plan_raw:
                plan = "intensive"
            else:
                plan = "professional"
            data[email] = {
                "plan": plan,
            }
    return data


def load_sharepoint_excel():
    """Load SharePoint Excel: Status, Intensive flag, Last Active Month."""
    from openpyxl import load_workbook
    candidates = [
        Path(SHAREPOINT_EXCEL_PATH),
        *glob.glob(str(DATA_DIR / "Babbel User List*")),
    ]
    filepath = next((f for f in candidates if Path(f).exists()), None)
    if not filepath:
        print("WARNUNG: SharePoint Excel nicht gefunden.")
        print(f"  Erwartet: {SHAREPOINT_EXCEL_PATH}")
        print(f"  Oder: daten/Babbel User List*.xlsx")
        return {}
    print(f"    Quelle: {filepath}")
    try:
        wb = load_workbook(filepath, read_only=True)
    except Exception as e:
        print(f"WARNUNG: SharePoint Excel konnte nicht gelesen werden: {e}")
        print("  (Datei eventuell von anderem Nutzer gesperrt oder OneDrive nicht gesynced)")
        return {}
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Spalten: [0]Name [1]E-mail [2]Entry [3]Professional [4]Intensive 2026
    #          [5]Credits Intensive(2026) [6]Credit Restock(2026)
    #          [7]Former Intensive(2025) [8]Credits Intensive(2025) [9]Credit Restock(2025)
    #          [10]Last Active Month [11]Status [12]Waitinglist [13]Warned [14]Comment
    data = {}
    for row in rows[1:]:
        email = (str(row[1]).strip().lower() if row[1] else "")
        if not email or email == "none":
            continue
        name = str(row[0]).strip() if row[0] else ""
        status = str(row[11]).strip() if row[11] else ""
        has_intensive_2026 = bool(row[4])
        has_professional = bool(row[3])
        last_active = None
        if row[10] and hasattr(row[10], "strftime"):
            last_active = row[10].strftime("%Y-%m-%d")
        comment = str(row[14]).strip() if row[14] else ""
        waitinglist = bool(row[12])
        warned = bool(row[13])

        data[email] = {
            "sp_name": name,
            "sp_status": status,
            "sp_has_professional": has_professional,
            "sp_has_intensive_2026": has_intensive_2026,
            "sp_last_active_month": last_active,
            "sp_comment": comment,
            "sp_waitinglist": waitinglist,
            "sp_warned": warned,
        }
    return data


# ---------------------------------------------------------------------------
# Consolidation & Scoring
# ---------------------------------------------------------------------------

def consolidate(learners, app_lessons, credits, memberships, sharepoint):
    """Join all data sources by email and compute inactivity score."""
    # Sammle alle bekannten E-Mails
    all_emails = set()
    all_emails.update(learners.keys())
    all_emails.update(memberships.keys())
    all_emails.update(sharepoint.keys())

    today = datetime.now()
    profiles = []

    for email in sorted(all_emails):
        lr = learners.get(email, {})
        al = app_lessons.get(email, {})
        cr = credits.get(email, {})
        mb = memberships.get(email, {})
        sp = sharepoint.get(email, {})

        # Basisinfo — Name aus verschiedenen Quellen zusammensetzen
        name = lr.get("name", "") or sp.get("sp_name", "")
        if not name:
            # Aus E-Mail ableiten: vorname.nachname@... -> Vorname Nachname
            local_part = email.split("@")[0]
            name = " ".join(part.capitalize() for part in local_part.replace("-", " ").split("."))

        plan = mb.get("plan", "unknown")
        status = sp.get("sp_status", "")

        # Letzte Lektion
        latest_lesson = al.get("latest_lesson_date", "")
        days_since_last_lesson = None
        if latest_lesson:
            try:
                last_dt = datetime.strptime(latest_lesson, "%Y-%m-%d")
                days_since_last_lesson = (today - last_dt).days
            except ValueError:
                pass

        # Credits
        credits_remaining = cr.get("credits_remaining", 0)

        # Active Days
        active_days = lr.get("active_days", None)
        learning_minutes = lr.get("learning_minutes", 0)

        # --- Inactivity Score ---
        score = 0
        reasons = []

        # Letzte Lektion > 12 Monate oder nie
        if days_since_last_lesson is None and active_days is not None and active_days == 0:
            score += 35
            reasons.append("Nie eine Lektion gemacht")
        elif days_since_last_lesson is not None and days_since_last_lesson > 365:
            score += 35
            reasons.append(f"Letzte Lektion vor {days_since_last_lesson} Tagen")
        elif days_since_last_lesson is not None and days_since_last_lesson > 180:
            score += 15
            reasons.append(f"Letzte Lektion vor {days_since_last_lesson} Tagen")

        # 0 Active Days in 2 Jahren
        if active_days is not None and active_days == 0:
            score += 25
            reasons.append("0 aktive Tage (2J)")
        elif active_days is not None and active_days <= 2:
            score += 10
            reasons.append(f"Nur {active_days} aktive Tage (2J)")

        # Keine remaining Credits
        if email in credits and credits_remaining == 0:
            score += 10
            reasons.append("0 Credits remaining")
        elif email not in credits and plan == "professional":
            score += 10
            reasons.append("Kein Intensive")

        # Nur Professional (kein Intensive-Plan)
        if plan == "professional":
            score += 10
            reasons.append("Nur Professional")

        # 0 Learning Minutes
        if learning_minutes == 0 and active_days is not None:
            score += 5
            reasons.append("0 Lernminuten")

        # --- Overrides ---
        if credits_remaining > 0:
            score = min(score, 50)
            reasons = [f"Hat {credits_remaining} Credits remaining"] + reasons

        if status == "Deleted":
            score = -1  # Markierung: uebersprungen

        if status == "Invited":
            score = 0
            reasons = ["Neu eingeladen"]

        # Admin-Override
        if email in ADMIN_EMAILS:
            score = 0
            reasons = ["Admin-Account (vom Scoring ausgeschlossen)"]

        # Empfehlung
        if score >= 80:
            recommendation = "LOESCHEN"
        elif score >= 50:
            recommendation = "PRUEFEN"
        elif score == -1:
            recommendation = "GELOESCHT"
        else:
            recommendation = "BEHALTEN"

        profile = {
            "email": email,
            "name": name,
            "plan": plan,
            "status": status,
            "active_days": active_days,
            "learning_minutes": learning_minutes,
            "latest_lesson_date": latest_lesson or "nie",
            "days_since_last_lesson": days_since_last_lesson,
            "credits_remaining": credits_remaining,
            "score": score,
            "recommendation": recommendation,
            "reasons": reasons,
            "in_babbel": email in memberships,
            "in_sharepoint": email in sharepoint,
        }
        profiles.append(profile)

    # Sortiere nach Score (hoechster zuerst)
    profiles.sort(key=lambda p: (-p["score"], p["name"]))
    return profiles


# ---------------------------------------------------------------------------
# Discrepancy Detection
# ---------------------------------------------------------------------------

def find_discrepancies(profiles):
    """Find differences between Babbel and SharePoint data."""
    in_babbel_not_sp = []
    in_sp_not_babbel = []
    status_mismatch = []

    for p in profiles:
        if p["in_babbel"] and not p["in_sharepoint"]:
            in_babbel_not_sp.append(p)
        elif p["in_sharepoint"] and not p["in_babbel"] and p["status"] == "Active":
            in_sp_not_babbel.append(p)
        elif p["in_babbel"] and p["in_sharepoint"] and p["status"] == "Deleted":
            status_mismatch.append(p)

    return in_babbel_not_sp, in_sp_not_babbel, status_mismatch


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

def print_discrepancies(in_babbel_not_sp, in_sp_not_babbel, status_mismatch):
    """Print discrepancy report."""
    print("\n" + "=" * 70)
    print("DISKREPANZEN: SharePoint vs. Babbel")
    print("=" * 70)

    if in_babbel_not_sp:
        print(f"\n  In Babbel aber NICHT in SharePoint-Excel ({len(in_babbel_not_sp)}):")
        for p in in_babbel_not_sp:
            print(f"    - {p['name']} ({p['email']})")
    else:
        print("\n  Alle Babbel-Nutzer sind in der SharePoint-Excel vorhanden.")

    if in_sp_not_babbel:
        print(f"\n  In SharePoint (Active) aber NICHT in Babbel ({len(in_sp_not_babbel)}):")
        for p in in_sp_not_babbel:
            print(f"    - {p['name']} ({p['email']})")
    else:
        print("\n  Alle aktiven SharePoint-Nutzer sind in Babbel vorhanden.")

    if status_mismatch:
        print(f"\n  In SharePoint als 'Deleted' markiert, aber in Babbel noch aktiv ({len(status_mismatch)}):")
        for p in status_mismatch:
            print(f"    - {p['name']} ({p['email']})")
    else:
        print("\n  Keine Status-Widersprueche gefunden.")


def print_recommendations(profiles, context=None):
    """Print scored recommendations."""
    # Filter by context
    if context == "professional":
        filtered = [p for p in profiles if p["plan"] == "professional"]
    elif context == "intensive":
        filtered = [p for p in profiles if p["plan"] == "intensive"]
    else:
        filtered = profiles

    # Kategorisieren
    to_delete = [p for p in filtered if p["recommendation"] == "LOESCHEN"]
    to_check = [p for p in filtered if p["recommendation"] == "PRUEFEN"]
    to_keep = [p for p in filtered if p["recommendation"] == "BEHALTEN"]
    deleted = [p for p in filtered if p["recommendation"] == "GELOESCHT"]

    header = f"\n{'=' * 70}\nEMPFEHLUNGEN"
    if context:
        header += f" (Kontext: {context.upper()})"
    header += f"\n{'=' * 70}"
    print(header)

    if to_delete:
        print(f"\n  LOESCHEN EMPFOHLEN ({len(to_delete)} Nutzer):")
        print(f"  {'-' * 90}")
        for i, p in enumerate(to_delete, 1):
            lesson = p["latest_lesson_date"][:10] if p["latest_lesson_date"] != "nie" else "nie"
            all_reasons = " | ".join(p["reasons"])
            print(f"  {i}. {p['name']} ({p['email']})")
            print(f"     Plan: {p['plan']} | Letzte Lektion: {lesson} | Score: {p['score']}")
            print(f"     Gruende: {all_reasons}")
            print()

    if to_check:
        print(f"\n  PRUEFEN ({len(to_check)} Nutzer):")
        print(f"  {'-' * 90}")
        for i, p in enumerate(to_check, 1):
            lesson = p["latest_lesson_date"][:10] if p["latest_lesson_date"] != "nie" else "nie"
            all_reasons = " | ".join(p["reasons"])
            print(f"  {i}. {p['name']} ({p['email']})")
            print(f"     Plan: {p['plan']} | Letzte Lektion: {lesson} | Score: {p['score']}")
            print(f"     Gruende: {all_reasons}")
            print()

    print(f"\n  BEHALTEN: {len(to_keep)} Nutzer | BEREITS GELOESCHT: {len(deleted)} Nutzer")

    # Zusammenfassung
    active_count = len(to_delete) + len(to_check) + len(to_keep)
    print(f"\n{'=' * 70}")
    print(f"ZUSAMMENFASSUNG")
    print(f"  Aktive Nutzer: {active_count} / 100 Slots")
    print(f"  Freie Slots: {100 - active_count}")
    print(f"  Loeschen empfohlen: {len(to_delete)}")
    print(f"  Zu pruefen: {len(to_check)}")
    print(f"  Behalten: {len(to_keep)}")
    print(f"{'=' * 70}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Babbel: Konsolidierte Nutzeranalyse")
    parser.add_argument("--context", choices=["professional", "intensive"],
                        help="Zeige nur Empfehlungen fuer diesen Plan-Typ")
    args = parser.parse_args()

    print("Lade Datenquellen...")
    learners = load_learners()
    print(f"  Learners: {len(learners)} Nutzer")
    app_lessons = load_app_lessons()
    print(f"  App-Lessons: {len(app_lessons)} Nutzer")
    credits = load_intensive_credits()
    print(f"  Intensive Credits: {len(credits)} Nutzer")
    memberships = load_memberships()
    print(f"  Memberships: {len(memberships)} Nutzer")
    sharepoint = load_sharepoint_excel()
    print(f"  SharePoint Excel: {len(sharepoint)} Nutzer")

    print("\nKonsolidiere Daten...")
    profiles = consolidate(learners, app_lessons, credits, memberships, sharepoint)
    print(f"  {len(profiles)} Nutzerprofile erstellt.")

    # Diskrepanzen
    in_babbel_not_sp, in_sp_not_babbel, status_mismatch = find_discrepancies(profiles)
    print_discrepancies(in_babbel_not_sp, in_sp_not_babbel, status_mismatch)

    # Empfehlungen
    print_recommendations(profiles, context=args.context)

    # JSON speichern
    EXPORT_DIR.mkdir(exist_ok=True)
    output_path = EXPORT_DIR / "consolidated_profiles.json"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(profiles, f, ensure_ascii=False, indent=2, default=str)
    print(f"\nJSON gespeichert: {output_path}")


if __name__ == "__main__":
    main()
